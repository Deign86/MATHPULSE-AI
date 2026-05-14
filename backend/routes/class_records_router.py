"""
Class Records Template Generator & Import Router.

GET  /api/templates/class-records  -> Download Excel template (public)
POST /api/class-records/upload     -> Upload & parse filled template (teacher only)
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Form, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse

from services.inference_client import call_hf_chat_async
from services.wri_service import compute_wri

logger = logging.getLogger("mathpulse.class_records")

router = APIRouter(prefix="/api", tags=["class-records"])

# ─── Lazy Firebase ──────────────────────────────────────────────────────────

_firebase_firestore = None


def _get_firestore_client():
    global _firebase_firestore
    if _firebase_firestore is None:
        try:
            from firebase_admin import firestore as ff
            _firebase_firestore = ff
        except Exception:
            _firebase_firestore = None
    if _firebase_firestore is None:
        return None
    try:
        return _firebase_firestore.client()
    except Exception:
        return None


# ─── Transmutation Table ───────────────────────────────────────────────────

TRANSMUTATION_TABLE: List[Tuple[float, float]] = [
    (100.00, 100), (98.40, 99), (96.80, 98), (95.20, 97), (93.60, 96),
    (92.00, 95), (90.40, 94), (88.80, 93), (87.20, 92), (85.60, 91),
    (84.00, 90), (82.40, 89), (80.80, 88), (79.20, 87), (77.60, 86),
    (76.00, 85), (74.40, 84), (72.80, 83), (71.20, 82), (69.60, 81),
    (68.00, 80), (66.40, 79), (64.80, 78), (63.20, 77), (61.60, 76),
    (60.00, 75), (58.40, 74), (56.80, 73), (55.20, 72), (53.60, 71),
    (52.00, 70), (50.40, 69), (48.80, 68), (47.20, 67), (45.60, 66),
    (44.00, 65), (42.40, 64), (40.80, 63), (39.20, 62), (37.60, 61),
    (36.00, 60), (34.40, 59), (32.80, 58), (31.20, 57), (29.60, 56),
    (28.00, 55), (26.40, 54), (24.80, 53), (23.20, 52), (21.60, 51),
    (20.00, 50),
]


def transmute(initial_grade: float) -> float:
    for threshold, transmuted in TRANSMUTATION_TABLE:
        if initial_grade >= threshold:
            return transmuted
    return 60.0


# ─── Quarter / SY Helpers ─────────────────────────────────────────────────

def _current_quarter() -> str:
    month = datetime.now().month
    if 1 <= month <= 3:
        return "Q1"
    if 4 <= month <= 6:
        return "Q2"
    if 7 <= month <= 9:
        return "Q3"
    return "Q4"


def _current_school_year() -> str:
    now = datetime.now()
    if now.month >= 8:
        return f"{now.year}-{now.year + 1}"
    return f"{now.year - 1}-{now.year}"


def _safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


# ─── OpenPyXL Style Helpers ────────────────────────────────────────────────

def _get_thin_border():
    from openpyxl.styles.borders import Border, Side
    side = Side(style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def _get_center_align():
    from openpyxl.styles.alignment import Alignment
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


# ─── ENDPOINT 1: GET /api/templates/class-records ─────────────────────────

@router.get("/templates/class-records")
async def download_class_records_template(
    quarter: Optional[str] = Query(default=None),
    school_year: Optional[str] = Query(default=None),
    subject: Optional[str] = Query(default=None),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Record"

    q = (quarter or "").strip() or _current_quarter()
    sy = (school_year or "").strip() or _current_school_year()
    subj = (subject or "").strip() or ""

    headers = [
        "LRN", "Last Name", "First Name", "Sex",
        "WW1", "WW2", "WW3", "WW4", "WW5",
        "PT1", "PT2", "PT3", "PT4", "PT5",
        "QA_Score", "QA_Max", "Absences_Q", "Remarks",
    ]
    num_cols = len(headers)

    thin_border = _get_thin_border()
    center_align = _get_center_align()

    purple_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    purple_font = Font(bold=True, color="3B0764", size=11)
    light_gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    italic_font = Font(italic=True, size=10)
    blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    green_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")
    red_fill = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Row 1: Teacher Name + Section
    for col, val in [(1, "Teacher Name:"), (3, "Section:")]:
        c = ws.cell(row=1, column=col, value=val)
        c.font = italic_font
        c.fill = light_gray_fill
        c.border = thin_border
    for col in [2, 4]:
        c = ws.cell(row=1, column=col)
        c.fill = light_gray_fill
        c.border = thin_border

    # Row 2: Subject + Quarter + School Year
    for col, val in [(1, "Subject:"), (3, "Quarter:"), (5, "School Year:")]:
        c = ws.cell(row=2, column=col, value=val)
        c.font = italic_font
        c.fill = light_gray_fill
        c.border = thin_border
    for col, val in [(2, subj), (4, q), (6, sy)]:
        c = ws.cell(row=2, column=col, value=val)
        c.fill = light_gray_fill
        c.border = thin_border

    # Row 3: Max Score per WW + PT
    for col, val in [(1, "Max Score per WW:"), (3, "Max Score per PT:")]:
        c = ws.cell(row=3, column=col, value=val)
        c.font = italic_font
        c.fill = light_gray_fill
        c.border = thin_border
    for col in [2, 4]:
        c = ws.cell(row=3, column=col)
        c.fill = light_gray_fill
        c.border = thin_border

    # Row 4: Headers
    for i, header in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=header)
        c.font = purple_font
        c.fill = purple_fill
        c.alignment = center_align
        c.border = thin_border

    # Column-level fills for data rows
    col_fills: Dict[int, Any] = {}
    for ci in range(5, 10):
        col_fills[ci] = blue_fill
    for ci in range(10, 15):
        col_fills[ci] = green_fill
    for ci in range(15, 17):
        col_fills[ci] = yellow_fill
    col_fills[17] = red_fill

    unlocked = Protection(locked=False)

    # Data rows 5-54
    for row in range(5, 55):
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.alignment = center_align
            if col in col_fills:
                c.fill = col_fills[col]
            else:
                c.fill = alt_fill if row % 2 == 0 else white_fill
            c.protection = unlocked

    # Data validation: Sex column (D) dropdown
    sex_dv = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
    sex_dv.error = "Please select M or F"
    sex_dv.errorTitle = "Invalid Sex"
    ws.add_data_validation(sex_dv)
    sex_dv.add("D5:D54")

    # Score columns: whole number >= 0
    score_col_indices = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
    for col_idx in score_col_indices:
        letter = get_column_letter(col_idx)
        dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        dv.error = "Score must be a non-negative whole number"
        dv.errorTitle = "Invalid Score"
        ws.add_data_validation(dv)
        dv.add(f"{letter}5:{letter}54")

    # Column widths
    widths = {"A": 15, "B": 20, "C": 20, "D": 10,
              "E": 10, "F": 10, "G": 10, "H": 10, "I": 10,
              "J": 10, "K": 10, "L": 10, "M": 10, "N": 10,
              "O": 10, "P": 10, "Q": 12, "R": 30}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = "E5"
    ws.protection.sheet = True

    # ── Instructions sheet ──
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        ("HOW TO FILL OUT THE CLASS RECORD TEMPLATE", True),
        ("", False),
        ("1. Fill in your Teacher Name and Section in Row 1.", False),
        ("2. Fill in the Subject in Row 2 (if not pre-filled).", False),
        ("3. Enter the Max Score per Written Work (WW) and Max Score per Performance Task (PT) in Row 3.", False),
        ("4. For each student (Rows 5-54), provide:", False),
        ("   - LRN (Learner Reference Number)", False),
        ("   - Last Name, First Name, Sex (M or F)", False),
        ("   - WW1 through WW5: Written Work scores (up to 5 entries)", False),
        ("   - PT1 through PT5: Performance Task scores (up to 5 entries)", False),
        ("   - QA_Score and QA_Max: Quarterly Assessment score", False),
        ("   - Absences_Q: Number of absences for the quarter", False),
        ("   - Remarks: Any additional notes", False),
        ("", False),
        ("ABBREVIATIONS", False),
        ("   WW = Written Work (25% of Initial Grade)", False),
        ("   PT = Performance Task (50% of Initial Grade)", False),
        ("   QA = Quarterly Assessment (25% of Initial Grade)", False),
        ("", False),
        ("DEPED FORMULA", False),
        ("   Initial Grade = (WW Average x 25%) + (PT Average x 50%) + (QA Grade x 25%)", False),
        ("   Transmuted Grade = DepEd transmutation table", False),
        ("   Passing Grade = 75 (transmuted)", False),
        ("", False),
        ("IMPORTANT", True),
        ("   Do NOT rename or restructure columns.", False),
        ("   Do NOT delete rows above Row 4.", False),
        ("   Removing or renaming columns will cause upload errors.", False),
        ("   Only the first sheet (Class Record) will be processed.", False),
    ]
    for i, (text, is_bold) in enumerate(instructions, 1):
        c = ws2.cell(row=i, column=1, value=text)
        c.font = Font(bold=is_bold, size=14 if is_bold and i == 1 else 11)
    ws2.column_dimensions["A"].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"MathPulse_ClassRecord_Template_{q}_{sy}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── ENDPOINT 2: POST /api/class-records/upload ────────────────────────────

ALLOWED_UPLOAD_EXTS = {".xlsx", ".csv"}


@router.post("/class-records/upload")
async def upload_class_records(
    request: Request,
    file: UploadFile = File(...),
    subject: Optional[str] = Form(None),
    quarter: Optional[str] = Form(None),
    gradeLevel: Optional[str] = Form(None),
):
    user: Any = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user.role not in {"teacher", "admin"}:
        raise HTTPException(status_code=403, detail="Only teachers and admins can upload class records")

    teacher_uid = user.uid

    # ── File validation ──
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_UPLOAD_EXTS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type '{ext}'. Accepted: {', '.join(ALLOWED_UPLOAD_EXTS)}")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Empty file")

    # ── Parse file ──
    try:
        if ext == ".csv":
            metadata, students = _parse_csv(contents)
        else:
            metadata, students = _parse_xlsx(contents)
    except (ValueError, KeyError, IndexError) as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    # Override metadata with form field values if provided
    if subject:
        metadata["subject"] = subject
    if quarter:
        metadata["quarter"] = quarter
    if gradeLevel:
        metadata["section"] = gradeLevel

    if not students:
        raise HTTPException(status_code=400, detail="No student data found in file")

    max_ww = metadata.get("max_ww", 0)
    max_pt = metadata.get("max_pt", 0)
    if not max_ww or not max_pt:
        raise HTTPException(status_code=400, detail="Max Score per WW and Max Score per PT must be set in Row 3")

    # ── Process students ──
    processed: List[Dict[str, Any]] = []

    # Human-readable flag descriptions
    FLAG_LABELS = {
        "failing_grade": "Failing grade",
        "very_low_grade": "Very low grade",
        "low_ww_average": "Low written work average",
        "low_pt_average": "Low performance task average",
        "failed_qa": "Failed quarterly assessment",
        "high_absences": "High absences",
        "missing_scores": "Missing scores",
        "declining_ww": "Declining written work trend",
        "declining_pt": "Declining performance task trend",
        "remarks_flag": "Concerning remarks",
    }
    MAX_FLAGS = len(FLAG_LABELS)

    for s in students:
        lrn = _safe_str(s.get("lrn"))
        if not lrn:
            continue

        ww_scores = [s.get(f"ww{i}") for i in range(1, 6)]
        pt_scores = [s.get(f"pt{i}") for i in range(1, 6)]
        qa_score = _safe_float(s.get("qa_score"))
        qa_max = _safe_float(s.get("qa_max"))
        absences = _safe_float(s.get("absences")) or 0
        remarks = _safe_str(s.get("remarks"))

        ww_floats = [_safe_float(v) for v in ww_scores]
        pt_floats = [_safe_float(v) for v in pt_scores]

        ww_valid = [v for v in ww_floats if v is not None]
        pt_valid = [v for v in pt_floats if v is not None]

        num_ww = len(ww_valid)
        num_pt = len(pt_valid)

        ww_total = (sum(ww_valid) / (max_ww * num_ww)) * 100 if max_ww and num_ww > 0 else 0.0
        pt_total = (sum(pt_valid) / (max_pt * num_pt)) * 100 if max_pt and num_pt > 0 else 0.0
        qa_grade = (qa_score / qa_max) * 100 if qa_score is not None and qa_max and qa_max > 0 else 0.0
        initial_grade = (ww_total * 0.25) + (pt_total * 0.50) + (qa_grade * 0.25)
        transmuted_grade = transmute(initial_grade)

        # At-risk flags
        flags: List[str] = []
        if transmuted_grade < 75:
            flags.append("failing_grade")
        if transmuted_grade < 70:
            flags.append("very_low_grade")
        ww_pct = (sum(ww_valid) / (max_ww * num_ww)) * 100 if max_ww and num_ww > 0 else 100.0
        pt_pct = (sum(pt_valid) / (max_pt * num_pt)) * 100 if max_pt and num_pt > 0 else 100.0
        qa_pct = qa_grade
        if ww_pct < 60:
            flags.append("low_ww_average")
        if pt_pct < 60:
            flags.append("low_pt_average")
        if qa_pct < 60:
            flags.append("failed_qa")
        if absences >= 4:
            flags.append("high_absences")

        missing_scores = any(v is None for v in ww_floats + pt_floats) or qa_score is None
        if missing_scores:
            flags.append("missing_scores")

        if len(ww_valid) >= 2 and ww_valid[-1] is not None and ww_valid[-2] is not None:
            if ww_valid[-1] < ww_valid[-2]:
                flags.append("declining_ww")
        if len(pt_valid) >= 2 and pt_valid[-1] is not None and pt_valid[-2] is not None:
            if pt_valid[-1] < pt_valid[-2]:
                flags.append("declining_pt")

        concern_words = ["absent", "failing", "concern", "struggle", "difficulty", "needs help", "intervention"]
        if any(w in remarks.lower() for w in concern_words):
            flags.append("remarks_flag")

        risk_level: str = "high" if len(flags) >= 3 else ("medium" if len(flags) >= 1 else "safe")

        # Compute risk score as percentage of flags triggered
        risk_score = round((len(flags) / MAX_FLAGS) * 100, 1)
        top_factors = [FLAG_LABELS.get(f, f) for f in flags[:5]]

        first_name = _safe_str(s.get("first_name"))
        last_name = _safe_str(s.get("last_name"))
        full_name = f"{first_name} {last_name}".strip() or lrn

        student_data = {
            "lrn": lrn,
            "lastName": last_name,
            "firstName": first_name,
            "sex": _safe_str(s.get("sex")),
            "wwScores": ww_floats,
            "ptScores": pt_floats,
            "qaScore": qa_score,
            "qaMax": qa_max,
            "absences": absences,
            "remarks": remarks,
            "wwTotal": round(ww_total, 2),
            "ptTotal": round(pt_total, 2),
            "qaGrade": round(qa_grade, 2),
            "initialGrade": round(initial_grade, 2),
            "transmutedGrade": round(transmuted_grade, 2),
            "flags": flags,
            "riskLevel": risk_level,
            "riskScore": risk_score,
            "topFactors": top_factors,
        }
        processed.append(student_data)

    summary = {
        "totalStudents": len(processed),
        "atRiskCount": sum(1 for s in processed if s["riskLevel"] == "high"),
        "mediumRiskCount": sum(1 for s in processed if s["riskLevel"] == "medium"),
        "lowRiskCount": sum(1 for s in processed if s["riskLevel"] == "safe"),
    }

    # ── Persist to Firestore ──
    upload_id = uuid.uuid4().hex
    upload_persisted = _persist_upload(teacher_uid, upload_id, metadata, processed, summary)

    section_id = _safe_str(metadata.get("section", "unknown")).replace(" ", "_").lower() or "unknown"
    students_persisted = _persist_students(teacher_uid, section_id, processed, upload_id)

    persisted = upload_persisted and students_persisted

    # D1+D3: Trigger WRI recompute for each student + cross-reference LRN → user account
    _trigger_wri_recompute(teacher_uid, section_id, processed)

    student_list = []
    for s in processed:
        raw_level = s["riskLevel"]
        mapped_level: str = "high" if raw_level == "high" else ("medium" if raw_level == "medium" else "low")
        student_list.append({
            "name": f"{s.get('firstName', '')} {s.get('lastName', '')}".strip() or s.get("lrn", ""),
            "riskLevel": mapped_level,
            "riskScore": s.get("riskScore", 0),
            "topFactors": s.get("topFactors", []),
        })

    response: Dict[str, Any] = {
        "success": True,
        "message": f"Uploaded {summary['totalStudents']} student records. {summary['atRiskCount']} at-risk, {summary['mediumRiskCount']} medium-risk, {summary['lowRiskCount']} safe.",
        "uploadId": upload_id,
        "persisted": persisted,
        "metadata": {
            "className": metadata.get("section", ""),
            "subject": metadata.get("subject", ""),
            "quarter": metadata.get("quarter", ""),
            "schoolYear": metadata.get("school_year", ""),
        },
        "summary": summary,
        "students": student_list,
    }
    if not persisted:
        response["warning"] = "Data not saved — Firestore unavailable"
    return response


# ─── ENDPOINT 3: POST /api/class-records/intervention-plan ──────────────────

@router.post("/class-records/intervention-plan")
async def create_intervention_plan(request: Request):
    """Generate a 3-step intervention plan for an at-risk student."""
    user: Any = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user.role not in {"teacher", "admin"}:
        raise HTTPException(status_code=403, detail="Only teachers and admins can create intervention plans")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    lrn = body.get("lrn")
    subject = body.get("subject")
    quarter = body.get("quarter")
    risk_factors = body.get("riskFactors", [])

    if not lrn or not subject or not quarter:
        raise HTTPException(status_code=400, detail="Missing required fields: lrn, subject, quarter")

    prompt = f"""You are an experienced Filipino senior high school teacher following DepEd guidelines.

Student LRN: {lrn}
Subject: {subject}
Quarter: {quarter}
Identified Risk Factors: {', '.join(risk_factors) if risk_factors else 'None specified'}

Create a concise, actionable 3-step intervention plan for this at-risk student.
Each step should be specific, measurable, and appropriate for the DepEd context.

Format your response as:
PLAN: <brief overall plan description>
STRATEGIES:
1. <first strategy>
2. <second strategy>
3. <third strategy>"""

    try:
        response_text = await call_hf_chat_async(
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500,
            temperature=0.7,
            task_type="intervention_plan"
        )

        # Parse the response
        plan = ""
        strategies: List[str] = []

        lines = response_text.strip().split("\n")
        in_strategies = False

        for line in lines:
            line = line.strip()
            if line.upper().startswith("PLAN:"):
                plan = line[5:].strip()
            elif line.upper().startswith("STRATEGIES:"):
                in_strategies = True
            elif in_strategies and line:
                # Remove leading numbers and dots
                cleaned = re.sub(r"^\d+\.\s*", "", line)
                if cleaned:
                    strategies.append(cleaned)

        # Ensure we have at least 3 strategies
        while len(strategies) < 3:
            strategies.append("Continue monitoring student progress and adjust interventions as needed")

        return {
            "plan": plan or f"Intervention plan for {subject} - {quarter}",
            "strategies": strategies[:3]
        }
    except Exception as e:
        logger.error(f"Failed to generate intervention plan: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate intervention plan: {str(e)}")


# ─── ENDPOINT 4: GET /api/class-records ─────────────────────────────────────

@router.get("/class-records")
async def get_class_records(
    request: Request,
    limit: int = Query(20, ge=1, le=100),
    after: Optional[str] = Query(None),
):
    """Get paginated class record uploads for the authenticated teacher."""
    user: Any = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user.role not in {"teacher", "admin"}:
        raise HTTPException(status_code=403, detail="Only teachers and admins can view class records")

    teacher_uid = user.uid
    client = _get_firestore_client()

    if client is None:
        return {"uploads": [], "hasMore": False}

    try:
        uploads_ref = client.collection("classRecords").document(teacher_uid).collection("uploads")
        query = uploads_ref.order_by("uploadedAt", direction="DESCENDING").limit(limit + 1)

        if after:
            # Get the document to start after
            after_doc = uploads_ref.document(after).get()
            if after_doc.exists:
                query = query.start_after(after_doc)

        docs = query.stream()
        uploads: List[Dict[str, Any]] = []

        for doc in docs:
            data = doc.to_dict()
            if data:
                uploads.append({
                    "uploadId": data.get("uploadId", doc.id),
                    "uploadedAt": data.get("uploadedAt", ""),
                    "studentCount": data.get("studentCount", 0),
                    "summary": data.get("summary", {}),
                    "metadata": {
                        "section": data.get("section", ""),
                        "subject": data.get("subject", ""),
                        "quarter": data.get("quarter", ""),
                        "schoolYear": data.get("schoolYear", ""),
                    },
                })

        has_more = len(uploads) > limit
        if has_more:
            uploads = uploads[:limit]

        return {"uploads": uploads, "hasMore": has_more}
    except Exception as e:
        logger.error(f"Failed to fetch class records: {e}")
        return {"uploads": [], "hasMore": False}


# ─── ENDPOINT 5: GET /api/class-records/{uploadId}/students ──────────────────

@router.get("/class-records/{uploadId}/students")
async def get_upload_students(
    request: Request,
    uploadId: str,
    limit: int = Query(100, ge=1, le=500),
    after: Optional[str] = Query(None),
):
    user: Any = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user.role not in {"teacher", "admin"}:
        raise HTTPException(status_code=403, detail="Only teachers and admins can view student records")

    teacher_uid = user.uid
    client = _get_firestore_client()
    if client is None:
        raise HTTPException(status_code=503, detail="Firestore unavailable")

    upload_ref = client.collection("classRecords").document(teacher_uid).collection("uploads").document(uploadId)
    upload_snap = upload_ref.get()
    if not upload_snap.exists:
        raise HTTPException(status_code=404, detail="Upload not found")

    upload_data = upload_snap.to_dict()
    section_id = (upload_data.get("section") or "unknown").replace(" ", "_").lower()

    students_ref = client.collection("classRecords").document(teacher_uid).collection("sections").document(section_id).collection("students")
    query_ref = students_ref.limit(limit + 1)
    if after:
        after_doc = students_ref.document(after).get()
        if after_doc.exists:
            query_ref = query_ref.start_after(after_doc)

    docs = query_ref.stream()
    students: List[Dict[str, Any]] = []
    for d in docs:
        data = d.to_dict()
        if data:
            students.append(data)

    has_more = len(students) > limit
    if has_more:
        students = students[:limit]

    return {"uploadId": uploadId, "sectionId": section_id, "students": students, "hasMore": has_more}


# ─── ENDPOINT 6: POST /api/class-records/{uploadId}/ai-report ────────────────

@router.post("/class-records/{uploadId}/ai-report")
async def generate_ai_class_report(request: Request, uploadId: str):
    import json as _json

    user: Any = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user.role not in {"teacher", "admin"}:
        raise HTTPException(status_code=403, detail="Only teachers and admins can generate reports")

    teacher_uid = user.uid
    client = _get_firestore_client()
    if client is None:
        raise HTTPException(status_code=503, detail="Firestore unavailable")

    upload_ref = client.collection("classRecords").document(teacher_uid).collection("uploads").document(uploadId)
    upload_snap = upload_ref.get()
    if not upload_snap.exists:
        raise HTTPException(status_code=404, detail="Upload not found")

    upload_data = upload_snap.to_dict()
    section_id = (upload_data.get("section") or "unknown").replace(" ", "_").lower()
    metadata = upload_data.get("metadata", {})

    students_ref = client.collection("classRecords").document(teacher_uid).collection("sections").document(section_id).collection("students")
    docs = students_ref.stream()
    all_students = [d.to_dict() for d in docs if d.to_dict()]

    total = len(all_students)
    if total == 0:
        return {"error": "No students found", "classAverage": 0, "distribution": [], "atRiskPct": 0, "recommendations": []}

    grades = [s.get("transmutedGrade", 0) for s in all_students if s.get("transmutedGrade") is not None]
    class_avg = round(sum(grades) / len(grades), 2) if grades else 0

    excellent = sum(1 for g in grades if g >= 90)
    very_satisfactory = sum(1 for g in grades if 85 <= g < 90)
    satisfactory = sum(1 for g in grades if 80 <= g < 85)
    fairly_satisfactory = sum(1 for g in grades if 75 <= g < 80)
    did_not_meet = sum(1 for g in grades if g < 75)

    distribution = [
        {"label": "Excellent (90-100)", "count": excellent, "pct": round(excellent / total * 100, 1) if total else 0},
        {"label": "Very Satisfactory (85-89)", "count": very_satisfactory, "pct": round(very_satisfactory / total * 100, 1) if total else 0},
        {"label": "Satisfactory (80-84)", "count": satisfactory, "pct": round(satisfactory / total * 100, 1) if total else 0},
        {"label": "Fairly Satisfactory (75-79)", "count": fairly_satisfactory, "pct": round(fairly_satisfactory / total * 100, 1) if total else 0},
        {"label": "Did Not Meet (<75)", "count": did_not_meet, "pct": round(did_not_meet / total * 100, 1) if total else 0},
    ]
    at_risk_pct = round(did_not_meet / total * 100, 1) if total else 0

    prompt = (
        "You are a DepEd school analyst. Given this class grade distribution for "
        f"{metadata.get('section','')} {metadata.get('subject','')} Q{metadata.get('quarter','')}, "
        "write a brief Quarterly Assessment Report. "
        f"Class avg: {class_avg}. Distribution: Excellent={excellent}, VS={very_satisfactory}, "
        f"S={satisfactory}, FS={fairly_satisfactory}, DNME={did_not_meet}. "
        'Write 2 specific actionable recommendations. Return JSON: {"recommendations": ["rec1", "rec2"]}'
    )

    try:
        response_text = await call_hf_chat_async(
            messages=[{"role": "user", "content": prompt}],
            max_tokens=400, temperature=0.5, task_type="class_report"
        )
        try:
            parsed = _json.loads(response_text)
            recommendations = parsed.get("recommendations", [])
        except Exception:
            recommendations = [r.strip() for r in response_text.split('\n') if len(r.strip()) > 20][:2]
    except Exception as e:
        logger.error(f"AI class report failed: {e}")
        recommendations = []

    return {
        "classAverage": class_avg,
        "distribution": distribution,
        "atRiskPct": at_risk_pct,
        "recommendations": recommendations,
    }


# ─── Parsers ────────────────────────────────────────────────────────────────

def _parse_xlsx(contents: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")

    wb = load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("No active sheet found")

    rows = list(ws.iter_rows(min_row=1, max_row=54, values_only=True))
    if len(rows) < 4:
        raise ValueError("File must have at least 4 rows (metadata + headers)")

    # Metadata
    r1, r2, r3 = rows[0], rows[1], rows[2]
    teacher_name = _safe_str(r1[1] if len(r1) > 1 else "")
    section = _safe_str(r1[3] if len(r1) > 3 else "")
    subject = _safe_str(r2[1] if len(r2) > 1 else "")
    quarter = _safe_str(r2[3] if len(r2) > 3 else "")
    school_year = _safe_str(r2[5] if len(r2) > 5 else "")
    max_ww = _safe_float(r3[1] if len(r3) > 1 else 0) or 0
    max_pt = _safe_float(r3[3] if len(r3) > 3 else 0) or 0

    # Validate header row
    expected_headers = ["LRN", "Last Name", "First Name", "Sex",
                        "WW1", "WW2", "WW3", "WW4", "WW5",
                        "PT1", "PT2", "PT3", "PT4", "PT5",
                        "QA_Score", "QA_Max", "Absences_Q", "Remarks"]
    r4 = [_safe_str(v) for v in (rows[3] if len(rows) > 3 else [])]
    if not r4 or r4[0] != "LRN":
        raise ValueError("Template structure not recognized. Expected Row 4 to start with 'LRN'.")

    # Data rows
    students: List[Dict[str, Any]] = []
    for row in rows[4:]:
        if not row or not _safe_str(row[0] if len(row) > 0 else ""):
            continue
        lrn = _safe_str(row[0] if len(row) > 0 else "")
        students.append({
            "lrn": lrn,
            "last_name": _safe_str(row[1] if len(row) > 1 else ""),
            "first_name": _safe_str(row[2] if len(row) > 2 else ""),
            "sex": _safe_str(row[3] if len(row) > 3 else ""),
            "ww1": row[4] if len(row) > 4 else None,
            "ww2": row[5] if len(row) > 5 else None,
            "ww3": row[6] if len(row) > 6 else None,
            "ww4": row[7] if len(row) > 7 else None,
            "ww5": row[8] if len(row) > 8 else None,
            "pt1": row[9] if len(row) > 9 else None,
            "pt2": row[10] if len(row) > 10 else None,
            "pt3": row[11] if len(row) > 11 else None,
            "pt4": row[12] if len(row) > 12 else None,
            "pt5": row[13] if len(row) > 13 else None,
            "qa_score": row[14] if len(row) > 14 else None,
            "qa_max": row[15] if len(row) > 15 else None,
            "absences": row[16] if len(row) > 16 else 0,
            "remarks": _safe_str(row[17] if len(row) > 17 else ""),
        })

    metadata = {
        "teacher_name": teacher_name,
        "section": section,
        "subject": subject,
        "quarter": quarter,
        "school_year": school_year,
        "max_ww": max_ww,
        "max_pt": max_pt,
    }
    return metadata, students


def _parse_csv(contents: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []

    if not fieldnames or "LRN" not in fieldnames:
        raise ValueError("CSV must have an 'LRN' column")

    metadata: Dict[str, Any] = {
        "teacher_name": "",
        "section": "",
        "subject": "",
        "quarter": "",
        "school_year": "",
        "max_ww": 0,
        "max_pt": 0,
    }

    students: List[Dict[str, Any]] = []
    for row in reader:
        lrn = _safe_str(row.get("LRN", ""))
        if not lrn:
            continue
        students.append({
            "lrn": lrn,
            "last_name": _safe_str(row.get("Last Name", "")),
            "first_name": _safe_str(row.get("First Name", "")),
            "sex": _safe_str(row.get("Sex", "")),
            "ww1": _safe_float(row.get("WW1")),
            "ww2": _safe_float(row.get("WW2")),
            "ww3": _safe_float(row.get("WW3")),
            "ww4": _safe_float(row.get("WW4")),
            "ww5": _safe_float(row.get("WW5")),
            "pt1": _safe_float(row.get("PT1")),
            "pt2": _safe_float(row.get("PT2")),
            "pt3": _safe_float(row.get("PT3")),
            "pt4": _safe_float(row.get("PT4")),
            "pt5": _safe_float(row.get("PT5")),
            "qa_score": _safe_float(row.get("QA_Score")),
            "qa_max": _safe_float(row.get("QA_Max")),
            "absences": _safe_float(row.get("Absences_Q")) or 0,
            "remarks": _safe_str(row.get("Remarks", "")),
        })

    return metadata, students


# ─── Firestore Persistence ──────────────────────────────────────────────────

def _persist_upload(
    teacher_uid: str,
    upload_id: str,
    metadata: Dict[str, Any],
    students: List[Dict[str, Any]],
    summary: Dict[str, int],
) -> bool:
    client = _get_firestore_client()
    if client is None:
        logger.warning("Firestore unavailable; class records not persisted")
        return False

    try:
        upload_doc = {
            "uploadId": upload_id,
            "teacherId": teacher_uid,
            "teacherName": metadata.get("teacher_name", ""),
            "section": metadata.get("section", ""),
            "subject": metadata.get("subject", ""),
            "quarter": metadata.get("quarter", ""),
            "schoolYear": metadata.get("school_year", ""),
            "maxWw": int(metadata.get("max_ww", 0) or 0),
            "maxPt": int(metadata.get("max_pt", 0) or 0),
            "studentCount": len(students),
            "summary": summary,
            "uploadedAt": datetime.utcnow().isoformat(),
        }

        upload_ref = client.collection("classRecords").document(teacher_uid).collection("uploads").document(upload_id)
        upload_ref.set(upload_doc)

        logger.info(f"Upload record persisted: {teacher_uid}/{upload_id} ({len(students)} students)")
        return True
    except Exception as e:
        logger.error(f"Failed to persist upload record: {e}")
        return False


def _trigger_wri_recompute(
    teacher_uid: str,
    section_id: str,
    students: List[Dict[str, Any]],
) -> None:
    """After class records persist, recompute WRI for each student and cross-reference LRN → user account."""
    client = _get_firestore_client()
    if client is None:
        logger.warning("WRI recompute skipped: Firestore unavailable")
        return

    users_ref = client.collection("users")
    section_ref = client.collection("classRecords").document(teacher_uid).collection("sections").document(section_id)

    for s in students:
        lrn = (s.get("lrn") or "").strip()
        if not lrn:
            continue

        transmuted = s.get("transmutedGrade")
        if transmuted is None:
            continue

        try:
            result = compute_wri(d=None, g=transmuted, p=None)
            wri_value = result.get("wri")
            risk_status = result.get("risk_status", "pending_assessment")
        except Exception as e:
            logger.error(f"WRI computation failed for LRN {lrn}: {e}")
            continue

        try:
            student_ref = section_ref.collection("students").document(lrn)
            student_ref.set({
                "wriScore": wri_value,
                "wriRiskBand": risk_status,
                "wriComputedAt": datetime.utcnow().isoformat(),
            }, merge=True)
        except Exception as e:
            logger.error(f"Failed to write WRI to classRecords for LRN {lrn}: {e}")

        try:
            matched = users_ref.where("lrn", "==", lrn).limit(1).stream()
            for user_doc in matched:
                user_doc.reference.set({
                    "latestGrade": transmuted,
                    "wriExternalGrade": transmuted,
                    "wriScore": wri_value,
                    "wriRiskBand": risk_status,
                    "wriUpdatedAt": datetime.utcnow().isoformat(),
                }, merge=True)
                logger.info(f"LRN {lrn} matched to user {user_doc.id}: grade={transmuted}, WRI={wri_value}")
                break
        except Exception as e:
            logger.error(f"LRN cross-reference failed for LRN {lrn}: {e}")


def _persist_students(
    teacher_uid: str,
    section_id: str,
    students: List[Dict[str, Any]],
    upload_id: str,
) -> bool:
    client = _get_firestore_client()
    if client is None:
        return False

    batch = client.batch()
    section_ref = client.collection("classRecords").document(teacher_uid).collection("sections").document(section_id)
    count = 0

    try:
        for s in students:
            lrn = s.get("lrn", "")
            if not lrn:
                continue

            student_doc = {
                "lrn": lrn,
                "lastName": s.get("lastName", ""),
                "firstName": s.get("firstName", ""),
                "sex": s.get("sex", ""),
                "wwScores": s.get("wwScores", []),
                "ptScores": s.get("ptScores", []),
                "qaScore": s.get("qaScore"),
                "qaMax": s.get("qaMax"),
                "absences": s.get("absences", 0),
                "remarks": s.get("remarks", ""),
                "wwTotal": s.get("wwTotal", 0),
                "ptTotal": s.get("ptTotal", 0),
                "qaGrade": s.get("qaGrade", 0),
                "initialGrade": s.get("initialGrade", 0),
                "transmutedGrade": s.get("transmutedGrade", 0),
                "flags": s.get("flags", []),
                "riskLevel": s.get("riskLevel", "safe"),
                "uploadId": upload_id,
                "updatedAt": datetime.utcnow().isoformat(),
            }

            student_ref = section_ref.collection("students").document(lrn)
            batch.set(student_ref, student_doc, merge=True)
            count += 1

            if count % 500 == 0:
                try:
                    batch.commit()
                except Exception as e:
                    logger.error(f"Batch commit failed at {count} students: {e}")
                    return False
                batch = client.batch()

        if count % 500 != 0:
            try:
                batch.commit()
            except Exception as e:
                logger.error(f"Final batch commit failed: {e}")
                return False

        logger.info(f"Persisted {count} students to classRecords/{teacher_uid}/sections/{section_id}")
        return True
    except Exception as e:
        logger.error(f"Failed to persist students: {e}")
        return False
